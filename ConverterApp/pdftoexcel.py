import os
import sys
import uuid
import json
import logging
import zipfile
import unicodedata
from logging.handlers import RotatingFileHandler
from datetime import datetime
from functools import wraps
import threading

import pdfplumber
import fitz
import pandas as pd
from flask import Flask, render_template, request, send_file, jsonify
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill


from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

def clean_text(value):
    if isinstance(value, str):
        # Loại bỏ mọi ký tự điều khiển không hợp lệ với Excel
        value = ILLEGAL_CHARACTERS_RE.sub("", value)
    return value

# Determine base directory (works for both script and frozen exe)
if getattr(sys, 'frozen', False):
    BASE_DIR = sys._MEIPASS
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Initialize Flask app
app = Flask(__name__,
            template_folder=os.path.join(BASE_DIR, 'templates'),
            static_folder=os.path.join(BASE_DIR, 'static'))
app.secret_key = os.environ.get('SECRET_KEY', os.urandom(24))

# Configuration - use absolute paths
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
ALLOWED_EXTENSIONS = {'pdf'}
MAX_CONTENT_LENGTH = 16 * 1024 * 1024  # 16MB limit

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH

# Ensure upload folder exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Setup logging
LOG_FOLDER = os.path.join(BASE_DIR, 'logs')
os.makedirs(LOG_FOLDER, exist_ok=True)

logger = logging.getLogger('pdf_converter')
logger.setLevel(logging.INFO)
logger.propagate = False

file_handler = RotatingFileHandler(
    os.path.join(LOG_FOLDER, 'app.log'),
    maxBytes=10 * 1024 * 1024,  # 10MB
    backupCount=5
)
file_handler.setFormatter(logging.Formatter(
    '%(asctime)s - %(levelname)s - %(message)s'
))
logger.addHandler(file_handler)

console_handler = logging.StreamHandler()
console_handler.setFormatter(logging.Formatter(
    '%(asctime)s - %(levelname)s - %(message)s'
))
logger.addHandler(console_handler)

# Rate limiting storage (in production, use Redis)
rate_limit_storage = {}
rate_limit_lock = threading.Lock()
RATE_LIMIT_REQUESTS = 10  # requests per window
RATE_LIMIT_WINDOW = 60  # seconds

# Heartbeat tracking for auto-shutdown
last_heartbeat = datetime.now()
heartbeat_lock = threading.Lock()
HEARTBEAT_TIMEOUT = 30  # 30 seconds without heartbeat triggers shutdown
shutdown_event = threading.Event()


def get_client_ip():
    """Get client IP address."""
    if request.headers.get('X-Forwarded-For'):
        return request.headers.get('X-Forwarded-For').split(',')[0].strip()
    return request.remote_addr or '127.0.0.1'


def rate_limit(f):
    """Rate limiting decorator."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        client_ip = get_client_ip()
        current_time = datetime.now().timestamp()

        with rate_limit_lock:
            if client_ip not in rate_limit_storage:
                rate_limit_storage[client_ip] = []

            # Clean old requests
            rate_limit_storage[client_ip] = [
                t for t in rate_limit_storage[client_ip]
                if current_time - t < RATE_LIMIT_WINDOW
            ]
            # Nếu IP không còn request nào trong window, xoá hẳn key để
            # tránh rate_limit_storage phình to vô hạn theo thời gian
            if not rate_limit_storage[client_ip]:
                del rate_limit_storage[client_ip]
                rate_limit_storage[client_ip] = []

            if len(rate_limit_storage[client_ip]) >= RATE_LIMIT_REQUESTS:
                logger.warning(f"Rate limit exceeded for IP: {client_ip}")
                return jsonify({
                    'error': 'Rate limit exceeded. Please try again later.',
                    'retry_after': RATE_LIMIT_WINDOW
                }), 429

            rate_limit_storage[client_ip].append(current_time)

        return f(*args, **kwargs)
    return decorated_function


def allowed_file(filename):
    """Check if file extension is allowed."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def _norm_text(value):
    """
    Chuẩn hóa Unicode về dạng NFC.
    Đây là lớp bảo hiểm cuối; nguồn sửa lỗi chính là việc đọc text bằng fitz
    thay vì để pdfminer (bên dưới pdfplumber) tự giải mã ToUnicode CMap.
    """
    if isinstance(value, str):
        return unicodedata.normalize('NFC', value)
    return value


def _extract_tables_with_fitz(pdf, fdoc):
    """
    Dùng pdfplumber CHỈ để dò cấu trúc bảng (tọa độ ô: page.find_tables()),
    còn chữ thực tế trong từng ô được lấy bằng fitz (PyMuPDF), vì fitz giải
    mã glyph -> Unicode ổn định hơn nhiều với font tiếng Việt so với
    pdfminer/pdfminer-based text extraction (nguồn gây lỗi kiểu
    "DIỆU" -> "DIẼU").

    Trả về danh sách các bảng, mỗi bảng là list các row, mỗi row là list
    các giá trị ô (raw list format, không phải DataFrame).
    """
    tables_per_page = []  # [(rows, bbox_list)] theo từng trang
    for page_num, page in enumerate(pdf.pages):
        fpage = fdoc[page_num]
        found_tables = page.find_tables()
        page_tables = []
        bboxes = []
        for table in found_tables:
            rows = []
            for row in table.rows:
                row_vals = []
                for cell in row.cells:
                    if cell is None:
                        row_vals.append("")
                        continue
                    rect = fitz.Rect(*cell)
                    cell_text = fpage.get_text("text", clip=rect).strip()
                    row_vals.append(_norm_text(cell_text))
                rows.append(row_vals)
            if rows:
                page_tables.append(rows)
                bboxes.append(table.bbox)
        tables_per_page.append((page_tables, bboxes))
    return tables_per_page


def extract_pdf_content(pdf_path):
    """Extract text and tables from PDF (dùng cho chế độ 'allText')."""
    logger.info(f"Extracting content from: {pdf_path}")
    text_lines = []
    table_data = []

    try:
        with pdfplumber.open(pdf_path) as pdf, fitz.open(pdf_path) as fdoc:
            tables_per_page = _extract_tables_with_fitz(pdf, fdoc)

            for page_num, page in enumerate(pdf.pages):
                fpage = fdoc[page_num]
                page_tables, table_bboxes = tables_per_page[page_num]
                table_data.extend(page_tables)

                # Lấy chữ ngoài bảng bằng fitz, bỏ qua block nằm trong vùng bảng
                for block in fpage.get_text("blocks"):
                    x0, y0, x1, y1, text = block[0], block[1], block[2], block[3], block[4]
                    inside_table = any(
                        x0 >= bx0 - 1 and y0 >= by0 - 1 and x1 <= bx1 + 1 and y1 <= by1 + 1
                        for bx0, by0, bx1, by1 in table_bboxes
                    )
                    if not inside_table:
                        cleaned = text.strip()
                        if cleaned:
                            text_lines.append(_norm_text(cleaned))

        text_content = "\n".join(text_lines)
        logger.info(f"Extracted {len(text_content)} chars and {len(table_data)} tables")
    except Exception as e:
        logger.error(f"Error extracting content from {pdf_path}: {str(e)}")
        raise

    return text_content, table_data


def extract_tables_from_pdf(pdf_path):
    """
    Extract tables from PDF (dùng cho chế độ 'tablesOnly').
    Trả về danh sách pandas DataFrame để tương thích với các hàm
    write_tables_to_excel / create_csv / create_json / process_pdf_job
    vốn đang dùng df.values, df.columns, df.iterrows(), df.to_dict(...).
    """
    logger.info(f"Extracting tables from: {pdf_path}")
    all_tables = []
    try:
        with pdfplumber.open(pdf_path) as pdf, fitz.open(pdf_path) as fdoc:
            tables_per_page = _extract_tables_with_fitz(pdf, fdoc)

            for page_tables, _ in tables_per_page:
                for rows in page_tables:
                    if not rows:
                        continue
                    header = [h if h else "" for h in rows[0]]
                    # Khử trùng tên cột trống/trùng để DataFrame không lỗi
                    seen = {}
                    clean_header = []
                    for h in header:
                        if h in seen:
                            seen[h] += 1
                            clean_header.append(f"{h}_{seen[h]}")
                        else:
                            seen[h] = 0
                            clean_header.append(h)
                    df = pd.DataFrame(rows[1:], columns=clean_header)
                    all_tables.append(df)

        logger.info(f"Extracted {len(all_tables)} tables from {pdf_path}")
    except Exception as e:
        logger.error(f"Error extracting tables from {pdf_path}: {str(e)}")
        raise
    return all_tables


def create_excel(text_content, table_data):
    """
    Create Excel workbook from text and table data.

    Lưu ý: extract_pdf_content() đã tự tách bạch text_content (chữ ngoài
    bảng) và table_data (chữ trong bảng) ngay từ khâu trích xuất bằng fitz,
    nên ở đây KHÔNG cần dò tìm dòng text nào "trùng khớp" với bảng như bản
    gốc của tác giả (cách đó dựa vào việc pdfminer trộn lẫn text và bảng
    vào chung một luồng). Chỉ cần ghi text trước, rồi ghi các bảng nối
    tiếp phía sau.
    """
    logger.info("Creating Excel workbook")
    wb = Workbook()
    ws = wb.active
    ws.title = "PDF Content"

    ws['A1'] = ""
    text_lines = text_content.split('\n') if text_content else []
    current_row = 2
    for line in text_lines:
        ws.cell(row=current_row, column=1, value=clean_text(line))
        current_row += 1

    if table_data:
        current_row += 1  # dòng trống ngăn cách text và bảng

    for table in table_data:
        for row_index, row_data in enumerate(table, start=current_row):
            for col_index, value in enumerate(row_data, start=2):  # bắt đầu từ cột B
                ws.cell(row=row_index, column=col_index, value=clean_text(value))
        current_row += len(table) + 2  # khoảng cách giữa các bảng

    # Apply font to maintain formatting
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name='Times New Roman', size=11)

    logger.info("Excel workbook created successfully")
    return wb


def write_tables_to_excel(tables, excel_path):
    """Write tables data to an Excel file."""
    logger.info(f"Writing {len(tables)} tables to Excel: {excel_path}")
    workbook = Workbook()

    for table_num, table in enumerate(tables, start=1):
        sheet = workbook.create_sheet(title=f'Table_{table_num}')

        title_font = Font(name='Times New Roman', size=11, bold=True)
        info_font = Font(name='Times New Roman', size=10)
        alignment = Alignment(wrap_text=True, vertical='center')
        header_fill = PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')

        df = pd.DataFrame(table.values, columns=[col.title() if col else "" for col in table.columns])

        for col_num, column_header in enumerate(df.columns, start=1):
            if column_header:
                cell = sheet.cell(row=1, column=col_num, value=column_header)
                cell.font = title_font
                cell.alignment = alignment
                cell.fill = header_fill
                sheet.column_dimensions[get_column_letter(col_num)].width = max(len(str(column_header)) + 2, 10)

        for row_num, (_, row) in enumerate(df.iterrows(), start=2):
            for col_num, value in enumerate(row, start=1):
                cell = sheet.cell(row=row_num, column=col_num, value=value)
                cell.font = info_font
                cell.alignment = alignment
                current_width = sheet.column_dimensions[get_column_letter(col_num)].width
                sheet.column_dimensions[get_column_letter(col_num)].width = max(
                    current_width, len(str(value)) + 2
                )

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            max_text_length = max(len(str(cell.value)) if cell.value else 0 for cell in row)
            sheet.row_dimensions[row[0].row].height = 35 + (max_text_length // 50) * 5

    workbook.remove(workbook.active)
    workbook.save(excel_path)
    logger.info(f"Excel file saved: {excel_path}")


def create_csv(text_content, table_data, output_path):
    """Create CSV from extracted data."""
    logger.info(f"Creating CSV: {output_path}")
    all_data = []

    if text_content:
        for line in text_content.split('\n'):
            if line.strip():
                all_data.append([line.strip()])

    for table in table_data:
        for row in table:
            if row:
                all_data.append([str(cell) if cell else '' for cell in row])

    df = pd.DataFrame(all_data)
    df.to_csv(output_path, index=False, header=False, encoding='utf-8-sig')
    logger.info(f"CSV file saved: {output_path}")


def create_json(text_content, table_data, output_path):
    """Create JSON from extracted data."""
    logger.info(f"Creating JSON: {output_path}")
    result = {
        'text': text_content.split('\n') if text_content else [],
        'tables': []
    }

    for i, table in enumerate(table_data):
        table_dict = {
            'table_number': i + 1,
            'rows': []
        }
        for row in table:
            if row:
                table_dict['rows'].append([str(cell) if cell else '' for cell in row])
        result['tables'].append(table_dict)

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    logger.info(f"JSON file saved: {output_path}")


def delete_files(*file_paths):
    """Delete files safely."""
    for path in file_paths:
        if path and os.path.exists(path):
            try:
                os.remove(path)
                logger.info(f"Deleted file: {path}")
            except PermissionError:
                logger.warning(f"Could not delete file (in use): {path}")
            except Exception as e:
                logger.error(f"Error deleting file {path}: {str(e)}")


@app.route('/')
def index():
    """Serve main page."""
    logger.info(f"Index page accessed from {get_client_ip()}")
    return render_template('pdftoexcel.html')


@app.route('/upload', methods=['POST'])
@rate_limit
def upload():
    """Handle file upload and start processing."""
    client_ip = get_client_ip()
    logger.info(f"Upload request from {client_ip}")

    if 'pdfFile' not in request.files:
        logger.warning(f"No file in request from {client_ip}")
        return jsonify({'error': 'No file provided'}), 400

    files = request.files.getlist('pdfFile')

    # Filter valid PDF files
    valid_files = [f for f in files if f.filename and allowed_file(f.filename)]

    if not valid_files:
        logger.warning(f"No valid PDF files from {client_ip}")
        return jsonify({'error': 'No valid PDF files provided'}), 400

    processing_option = request.form.get('processingOption', 'allText')
    output_format = request.form.get('outputFormat', 'xlsx')

    # Map output format to extension
    ext_map = {'xlsx': '.xlsx', 'csv': '.csv', 'json': '.json'}
    output_ext = ext_map.get(output_format, '.xlsx')

    try:
        output_files = []
        temp_pdf_files = []

        for file in valid_files:
            # Save PDF with unique prefix to avoid conflicts
            original_filename = secure_filename(file.filename)
            base_name = os.path.splitext(original_filename)[0]
            unique_id = str(uuid.uuid4())[:8]
            pdf_filename = f"{unique_id}_{original_filename}"
            pdf_filepath = os.path.join(app.config['UPLOAD_FOLDER'], pdf_filename)
            file.save(pdf_filepath)
            temp_pdf_files.append(pdf_filepath)
            logger.info(f"File saved: {pdf_filepath}")

            # Output filename uses original name
            output_filename = f"{base_name}{output_ext}"
            output_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{unique_id}_{output_filename}")

            if processing_option == 'allText':
                text_content, table_data = extract_pdf_content(pdf_filepath)

                # PDF dạng scan/ảnh (không có lớp text) sẽ cho ra text_content
                # và table_data đều rỗng. Trước đây code vẫn tạo ra 1 file
                # Excel/CSV/JSON hợp lệ nhưng trống trơn mà không báo gì, khiến
                # người dùng tưởng convert thành công. Giờ bỏ qua file này và
                # cảnh báo, giống hệt cách xử lý của chế độ 'tablesOnly'.
                if not text_content and not table_data:
                    logger.warning(
                        f"No extractable text/tables in {original_filename} "
                        "(có thể là PDF dạng scan/ảnh, cần OCR)"
                    )
                    continue

                if output_format == 'xlsx':
                    excel_file = create_excel(text_content, table_data)
                    excel_file.save(output_path)
                elif output_format == 'csv':
                    create_csv(text_content, table_data, output_path)
                elif output_format == 'json':
                    create_json(text_content, table_data, output_path)

            elif processing_option == 'tablesOnly':
                tables = extract_tables_from_pdf(pdf_filepath)
                if tables:
                    if output_format == 'xlsx':
                        write_tables_to_excel(tables, output_path)
                    elif output_format == 'csv':
                        all_data = []
                        for table in tables:
                            df = pd.DataFrame(table.values, columns=table.columns)
                            all_data.append(df)
                        if all_data:
                            combined = pd.concat(all_data, ignore_index=True)
                            combined.to_csv(output_path, index=False, encoding='utf-8-sig')
                    elif output_format == 'json':
                        result = {'tables': []}
                        for i, table in enumerate(tables):
                            df = pd.DataFrame(table.values, columns=table.columns)
                            result['tables'].append({
                                'table_number': i + 1,
                                'columns': list(table.columns),
                                'data': df.to_dict('records')
                            })
                        with open(output_path, 'w', encoding='utf-8') as f:
                            json.dump(result, f, ensure_ascii=False, indent=2)
                else:
                    logger.warning(f"No tables found in {original_filename}")
                    continue

            if os.path.exists(output_path):
                output_files.append((output_path, output_filename))

        # Clean up PDF files
        delete_files(*temp_pdf_files)

        if not output_files:
            return jsonify({'error': 'No files were converted successfully'}), 400

        # Single file - return directly with original name
        if len(output_files) == 1:
            output_path, output_filename = output_files[0]
            response = send_file(
                output_path,
                as_attachment=True,
                download_name=output_filename
            )

            @response.call_on_close
            def cleanup():
                delete_files(output_path)

            return response

        # Multiple files - create ZIP
        zip_filename = f"converted_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        zip_path = os.path.join(app.config['UPLOAD_FOLDER'], zip_filename)

        # Nếu nhiều PDF có cùng tên gốc, output_filename sẽ trùng nhau và
        # file sau sẽ ghi đè file trước trong ZIP. Thêm hậu tố (2), (3)...
        # cho các tên bị trùng để không mất dữ liệu.
        used_names = {}
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for output_path, output_filename in output_files:
                if output_filename in used_names:
                    used_names[output_filename] += 1
                    name, ext = os.path.splitext(output_filename)
                    zip_entry_name = f"{name} ({used_names[output_filename]}){ext}"
                else:
                    used_names[output_filename] = 1
                    zip_entry_name = output_filename
                zipf.write(output_path, zip_entry_name)

        # Clean up individual output files
        for output_path, _ in output_files:
            delete_files(output_path)

        logger.info(f"Created ZIP with {len(output_files)} files: {zip_path}")

        response = send_file(zip_path, as_attachment=True, download_name=zip_filename)

        @response.call_on_close
        def cleanup():
            delete_files(zip_path)

        return response

    except Exception as e:
        logger.error(f"Error processing upload from {client_ip}: {str(e)}")
        return jsonify({'error': f'Processing error: {str(e)}'}), 500


@app.route('/health')
def health_check():
    """Health check endpoint."""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat()
    })


@app.route('/heartbeat', methods=['POST'])
def heartbeat():
    """Receive heartbeat from browser to keep server alive."""
    global last_heartbeat
    with heartbeat_lock:
        last_heartbeat = datetime.now()
    return jsonify({'status': 'ok'})


@app.route('/shutdown', methods=['POST'])
def shutdown():
    """Graceful shutdown endpoint (called when browser closes)."""
    logger.info("Shutdown requested via endpoint")
    shutdown_event.set()
    return jsonify({'status': 'shutting_down'})


def heartbeat_monitor():
    """Monitor heartbeat and shutdown server if no heartbeat received."""
    logger.info(f"Heartbeat monitor started (timeout: {HEARTBEAT_TIMEOUT}s)")

    while not shutdown_event.is_set():
        shutdown_event.wait(timeout=10)  # Check every 10 seconds

        with heartbeat_lock:
            elapsed = (datetime.now() - last_heartbeat).total_seconds()

        if elapsed > HEARTBEAT_TIMEOUT:
            logger.info(f"No heartbeat for {elapsed:.0f}s - initiating shutdown")
            shutdown_event.set()
            break

    logger.info("Heartbeat monitor: triggering server shutdown")
    os._exit(0)


def start_heartbeat_monitor():
    """Start the heartbeat monitor thread."""
    monitor_thread = threading.Thread(target=heartbeat_monitor, daemon=True)
    monitor_thread.start()
    return monitor_thread


@app.errorhandler(413)
def request_entity_too_large(error):
    """Handle file too large error."""
    logger.warning(f"File too large from {get_client_ip()}")
    return jsonify({
        'error': f'File too large. Maximum size is {MAX_CONTENT_LENGTH // (1024 * 1024)}MB'
    }), 413


@app.errorhandler(500)
def internal_error(error):
    import traceback
    traceback.print_exc()
    return jsonify({
        'error': str(error)
    }), 500

@app.errorhandler(Exception)
def handle_exception(e):
    from werkzeug.exceptions import HTTPException
    # Không nuốt các lỗi HTTP bình thường của Flask/werkzeug (404, 405...),
    # nếu không mọi request không khớp route (vd. GET /favicon.ico) sẽ bị
    # biến thành 500 thay vì giữ đúng mã lỗi gốc, và lỗi thật sẽ khó thấy hơn.
    if isinstance(e, HTTPException):
        return e

    import traceback
    logger.error(f"Unhandled exception: {str(e)}")
    traceback.print_exc()

    return jsonify({
        'error': str(e)
    }), 500


if __name__ == "__main__":
    debug = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    logger.info(f"Starting PDF to Excel Converter (debug={debug})")
    # QUAN TRỌNG: phải gọi hàm này thì cơ chế tự tắt server khi mất heartbeat
    # (trình duyệt đóng/mất kết nối > HEARTBEAT_TIMEOUT giây) mới thực sự
    # chạy. Trước đây hàm start_heartbeat_monitor() chỉ được ĐỊNH NGHĨA mà
    # không hề được gọi ở đâu, nên server luôn chạy mãi cho tới khi tắt
    # cửa sổ console/nhấn Ctrl+C, kể cả khi đã đóng trình duyệt.
    start_heartbeat_monitor()
    app.run(host='0.0.0.0', port=5000, debug=debug, threaded=True)